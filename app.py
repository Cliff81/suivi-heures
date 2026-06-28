#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Suivi des heures - Version web (Flask).

Reproduit l'application bureau (Saisie, Recapitulatif, Recap global,
Conversion, Salaries) dans un navigateur, avec comptes utilisateurs.
La logique metier et la base SQLite sont partagees via core.py.

Lancement :  python3 app.py   puis ouvrir http://127.0.0.1:5000
"""

import io
import os
import secrets
import datetime
import functools

from flask import (Flask, g, session, request, redirect, url_for,
                   render_template, flash, send_file, abort)
from werkzeug.security import generate_password_hash, check_password_hash

import core
from core import (DB, DB_PATH, MOIS_NOMS, JOURS_NOMS, SEUIL_ACCORD,
                  MAX_JOUR, MAX_SEMAINE, nom_complet, fmt_hm, fmt_hm_signe,
                  fmt_centieme, fmt_km, parse_heure, heures_travaillees,
                  totaux_semaine, totaux_mois, semaines_du_mois,
                  verifier_conformite)

app = Flask(__name__)

# Cle de session : priorite a la variable d'environnement SECRET_KEY (recommandee
# en ligne, ou le systeme de fichiers peut etre ephemere), sinon fichier local.
_secret_file = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            ".secret_key")
if os.environ.get("SECRET_KEY"):
    app.secret_key = os.environ["SECRET_KEY"]
elif os.path.exists(_secret_file):
    with open(_secret_file) as f:
        app.secret_key = f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    try:
        with open(_secret_file, "w") as f:
            f.write(app.secret_key)
    except OSError:
        pass

# Cookies de session securises derriere HTTPS (hebergement en ligne).
if os.environ.get("HTTPS_ONLY") == "1":
    app.config.update(SESSION_COOKIE_SECURE=True,
                      SESSION_COOKIE_HTTPONLY=True,
                      SESSION_COOKIE_SAMESITE="Lax")

# Compte par defaut cree si aucun utilisateur n'existe (a changer ensuite).
ADMIN_DEFAUT = ("admin", "admin")


# --------------------------------------------------------------------------
# Base de donnees par requete
# --------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = DB(DB_PATH)
        # amorcage du compte admin par defaut
        if not g.db.users():
            g.db.ajouter_user(ADMIN_DEFAUT[0],
                              generate_password_hash(ADMIN_DEFAUT[1]),
                              is_admin=1)
    return g.db


@app.teardown_appcontext
def _close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# --------------------------------------------------------------------------
# Authentification
# --------------------------------------------------------------------------
def login_required(view):
    @functools.wraps(view)
    def wrapped(*a, **kw):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        return view(*a, **kw)
    return wrapped


def admin_required(view):
    @functools.wraps(view)
    def wrapped(*a, **kw):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        if not session.get("is_admin"):
            # un simple utilisateur est renvoye vers son espace de saisie
            return redirect(url_for("espace"))
        return view(*a, **kw)
    return wrapped


@app.context_processor
def _inject():
    return dict(app_user=session.get("user"),
                app_is_admin=session.get("is_admin"),
                MOIS_NOMS=MOIS_NOMS, annee_courante=datetime.date.today().year)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        db = get_db()
        login_ = request.form.get("login", "").strip()
        pwd = request.form.get("password", "")
        u = db.get_user(login_)
        if u and check_password_hash(u["pwd_hash"], pwd):
            session.clear()
            session["user"] = u["login"]
            session["is_admin"] = bool(u["is_admin"])
            session["salarie_id"] = u["salarie_id"]
            accueil = url_for("saisie") if u["is_admin"] else url_for("espace")
            nxt = request.args.get("next") or accueil
            return redirect(nxt)
        flash("Identifiant ou mot de passe incorrect.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# --------------------------------------------------------------------------
# Outils communs
# --------------------------------------------------------------------------
def _annee_param(defaut=None):
    try:
        return int(request.args.get("annee", defaut or datetime.date.today().year))
    except (TypeError, ValueError):
        return datetime.date.today().year


def _sid_param(db):
    sals = db.salaries()
    if not sals:
        return None, sals
    try:
        sid = int(request.args.get("sid", sals[0]["id"]))
    except (TypeError, ValueError):
        sid = sals[0]["id"]
    if not any(s["id"] == sid for s in sals):
        sid = sals[0]["id"]
    return sid, sals


# --------------------------------------------------------------------------
# Accueil
# --------------------------------------------------------------------------
@app.route("/")
@login_required
def index():
    if session.get("is_admin"):
        return redirect(url_for("saisie"))
    return redirect(url_for("espace"))


# --------------------------------------------------------------------------
# Saisie
# --------------------------------------------------------------------------
@app.route("/saisie", methods=["GET", "POST"])
@admin_required
def saisie():
    db = get_db()
    sid, sals = _sid_param(db)
    annee = _annee_param()
    try:
        mois = int(request.args.get("mois", datetime.date.today().month))
    except (TypeError, ValueError):
        mois = datetime.date.today().month
    mois = min(12, max(1, mois))

    if sid is None:
        return render_template("saisie.html", sals=[], sal=None,
                               annee=annee, mois=mois)

    if request.method == "POST":
        _enregistrer_saisie(db, sid, annee, mois)
        flash("Enregistré.", "ok")
        return redirect(url_for("saisie", sid=sid, annee=annee, mois=mois))

    sal = db.salarie(sid)
    seuil = db.duree_hebdo(sid)
    semaines = semaines_du_mois(annee, mois)
    debut = semaines[0][0].isoformat()
    fin = semaines[-1][6].isoformat()
    data_map = db.jours_intervalle(sid, debut, fin)

    today = datetime.date.today()
    semaines_vue = []
    for num, sem in enumerate(semaines, start=1):
        jours = []
        for d in sem:
            iso = d.isoformat()
            hors_mois = (d.month != mois or d.year != annee)
            j = data_map.get(iso) if not hors_mois else None

            def val(f, km=False):
                if j is None or j[f] is None:
                    return ""
                return fmt_km(j[f]) if km else fmt_hm(j[f])

            jours.append({
                "iso": iso, "jour_nom": JOURS_NOMS[d.weekday()],
                "date_fr": d.strftime("%d/%m/%Y"),
                "hors_mois": hors_mois, "weekend": d.weekday() >= 5,
                "amp_d": val("amp_d"), "amp_f": val("amp_f"),
                "pause_d": val("pause_d"), "pause_f": val("pause_f"),
                "km_jour": val("km_jour", km=True),
                "km_com": val("km_com", km=True),
            })
        semaines_vue.append({
            "num": num,
            "du": sem[0].strftime("%d/%m"), "au": sem[6].strftime("%d/%m"),
            "jours": jours,
        })

    return render_template(
        "saisie.html", sals=sals, sal=sal, sid=sid, annee=annee, mois=mois,
        seuil=seuil, seuil_min=seuil, semaines=semaines_vue,
        payees=fmt_hm(db.payees(sid, annee, mois)),
        pris=fmt_hm(db.conges_pris(sid, annee, mois)),
        solde=fmt_hm_signe(db.solde_conges(sid, annee, mois)),
        MAX_JOUR=MAX_JOUR, MAX_SEMAINE=MAX_SEMAINE)


def _enregistrer_saisie(db, sid, annee, mois):
    """Enregistre toutes les journees affichees + compteurs du mois."""
    semaines = semaines_du_mois(annee, mois)
    for sem in semaines:
        for d in sem:
            if d.month != mois or d.year != annee:
                continue
            iso = d.isoformat()
            vals = {}
            ok = True
            for f in ("amp_d", "amp_f", "pause_d", "pause_f"):
                try:
                    vals[f] = parse_heure(request.form.get(f"{iso}__{f}", ""))
                except ValueError:
                    vals[f] = None
                    ok = False
            vals["km_jour"] = _lire_km(request.form.get(f"{iso}__km_jour", ""))
            vals["km_com"] = _lire_km(request.form.get(f"{iso}__km_com", ""))
            if vals["km_jour"] is None and vals["km_com"] is None:
                vals["km_vide"] = None
            else:
                vals["km_vide"] = (vals["km_jour"] or 0) - (vals["km_com"] or 0)
            if ok:
                db.enregistrer_jour(sid, iso, vals)
    db.commit()
    try:
        db.set_payees(sid, annee, mois,
                      parse_heure(request.form.get("payees", "")) or 0)
    except ValueError:
        pass
    try:
        db.set_conges_pris(sid, annee, mois,
                           parse_heure(request.form.get("pris", "")) or 0)
    except ValueError:
        pass


def _lire_km(s):
    s = (s or "").strip().replace(",", ".")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


# --------------------------------------------------------------------------
# Recapitulatif (par salarie)
# --------------------------------------------------------------------------
RECAP_LIGNES = [
    ("total", "Total heures travaillées"),
    ("sup35", "Total heures supplémentaires (> 35h)"),
    ("accord39", "Dont rentrant dans l'accord (contractuel)"),
    ("sup39", "Au-delà du contractuel (acquis congés)"),
    ("payees", "Heures supplémentaires payées"),
    ("pris", "Heures de congés prises"),
    ("cumul", "SOLDE compteur de congés (cumulé)"),
]


@app.route("/recap")
@admin_required
def recap():
    db = get_db()
    sid, sals = _sid_param(db)
    annee = _annee_param()
    if sid is None:
        return render_template("recap.html", sals=[], sal=None, annee=annee)
    sal = db.salarie(sid)

    data = {k: [] for k, _ in RECAP_LIGNES}
    cumul = db.solde_conges(sid, annee - 1, 12)
    for mois in range(1, 13):
        t = totaux_mois(db, sid, annee, mois)
        pris = db.conges_pris(sid, annee, mois)
        data["total"].append(t["total"])
        data["sup35"].append(t["sup35"])
        data["accord39"].append(t["accord39"])
        data["sup39"].append(t["sup39"])
        data["payees"].append(db.payees(sid, annee, mois))
        data["pris"].append(pris)
        cumul += t["sup39"] - pris
        data["cumul"].append(cumul)

    lignes = []
    for key, libelle in RECAP_LIGNES:
        vals = data[key]
        if key == "cumul":
            cellules = [fmt_hm_signe(v) for v in vals]
            total = fmt_hm_signe(vals[-1]) if vals else "00:00"
            css = "conges"
        else:
            cellules = [fmt_hm(v) for v in vals]
            total = fmt_hm(sum(vals))
            css = "sup" if key in ("sup35", "sup39") else ""
        lignes.append({"libelle": libelle, "cellules": cellules,
                       "total": total, "css": css})

    conf = _conformite_messages(db, sid, annee)
    return render_template("recap.html", sals=sals, sal=sal, sid=sid,
                           annee=annee, lignes=lignes,
                           mois_court=[m[:4] for m in MOIS_NOMS], conf=conf)


def _conformite_messages(db, sid, annee):
    c = verifier_conformite(db, sid, annee)
    msgs = []
    ok_global = True

    def jour_fr(d):
        return datetime.date.fromisoformat(d).strftime("%d/%m")

    if c["jours_ampl"]:
        ok_global = False
        ex = ", ".join(f"{jour_fr(d)} ({fmt_hm(m)})" for d, m in c["jours_ampl"][:6])
        suite = "…" if len(c["jours_ampl"]) > 6 else ""
        msgs.append((False, f"{len(c['jours_ampl'])} jour(s) amplitude > 12h : {ex}{suite}"))
    else:
        msgs.append((True, "Aucune amplitude > 12h"))

    if c["jours_10h"]:
        ok_global = False
        ex = ", ".join(f"{jour_fr(d)} ({fmt_hm(m)})" for d, m in c["jours_10h"][:6])
        suite = "…" if len(c["jours_10h"]) > 6 else ""
        msgs.append((False, f"{len(c['jours_10h'])} jour(s) > 10h : {ex}{suite}"))
    else:
        msgs.append((True, "Aucune journée > 10h"))

    if c["semaines_48h"]:
        ok_global = False
        ex = ", ".join(f"sem. du {d.strftime('%d/%m')} ({fmt_hm(m)})"
                       for d, m in c["semaines_48h"][:6])
        suite = "…" if len(c["semaines_48h"]) > 6 else ""
        msgs.append((False, f"{len(c['semaines_48h'])} semaine(s) > 48h : {ex}{suite}"))
    else:
        msgs.append((True, "Aucune semaine > 48h"))

    if c["moy12_depassee"]:
        ok_global = False
        f = c["moy12_fenetre"]
        per = (f"{f[0].strftime('%d/%m/%y')} → {f[1].strftime('%d/%m/%y')}"
               if f else "")
        msgs.append((False, f"Moyenne sur 12 semaines > 44h : {fmt_hm(c['moy12_max'])} ({per})"))
    else:
        msgs.append((True, f"Moyenne sur 12 sem. ≤ 44h (max : {fmt_hm(c['moy12_max'])})"))

    return {"ok": ok_global, "msgs": msgs}


# --------------------------------------------------------------------------
# Recap global (tous les salaries)
# --------------------------------------------------------------------------
GLOBAL_COLS = [
    ("total", "Heures travaillées"),
    ("sup35", "Sup. > 35h"),
    ("accord39", "Dans l'accord"),
    ("sup39", "Au-delà accord (acquis)"),
    ("payees", "Heures payées"),
    ("solde", "Solde congés"),
]


@app.route("/global")
@admin_required
def recap_global():
    db = get_db()
    annee = _annee_param()
    lignes = []
    totaux = {k: 0 for k, _ in GLOBAL_COLS}
    for sal in db.salaries():
        acc = {k: 0 for k, _ in GLOBAL_COLS}
        for mois in range(1, 13):
            t = totaux_mois(db, sal["id"], annee, mois)
            acc["total"] += t["total"]
            acc["sup35"] += t["sup35"]
            acc["accord39"] += t["accord39"]
            acc["sup39"] += t["sup39"]
            acc["payees"] += db.payees(sal["id"], annee, mois)
        acc["solde"] = db.solde_conges(sal["id"], annee, 12)
        for k in totaux:
            totaux[k] += acc[k]
        lignes.append({
            "nom": nom_complet(sal),
            "cellules": [(fmt_hm_signe(acc[k]) if k == "solde" else fmt_hm(acc[k]))
                         for k, _ in GLOBAL_COLS],
        })
    total_row = [(fmt_hm_signe(totaux[k]) if k == "solde" else fmt_hm(totaux[k]))
                 for k, _ in GLOBAL_COLS]
    bandeau = {
        "total": fmt_hm(totaux["total"]),
        "total_cent": fmt_centieme(totaux["total"]),
        "sup35": fmt_hm(totaux["sup35"]),
        "sup39": fmt_hm(totaux["sup39"]),
    }
    return render_template("global.html", annee=annee, lignes=lignes,
                           entetes=[lib for _, lib in GLOBAL_COLS],
                           total_row=total_row, bandeau=bandeau)


# --------------------------------------------------------------------------
# Conversion
# --------------------------------------------------------------------------
@app.route("/conversion")
@admin_required
def conversion():
    table = [(m, f"{m/60:.2f}".replace(".", ",")) for m in range(5, 61, 5)]
    return render_template("conversion.html", table=table)


# --------------------------------------------------------------------------
# Relevé mensuel des heures travaillées (par salarié, sans les pauses)
# --------------------------------------------------------------------------
def _mois_param():
    try:
        m = int(request.args.get("mois", datetime.date.today().month))
    except (TypeError, ValueError):
        m = datetime.date.today().month
    return min(12, max(1, m))


def _sids_selectionnes(db):
    """Liste des ids de salaries coches (parametres 'sid' multiples).
    Si aucun, tous les salaries."""
    sals = db.salaries()
    demandes = request.args.getlist("sid")
    if not demandes:
        return [s["id"] for s in sals]
    ids = set()
    for v in demandes:
        try:
            ids.add(int(v))
        except (TypeError, ValueError):
            pass
    return [s["id"] for s in sals if s["id"] in ids]


def _releve_data(db, sid, annee, mois):
    """Construit le releve d'un salarie : semaines -> jours travailles
    (heures nettes, pauses deduites), sous-totaux et total du mois.
    Les week-ends non travailles sont masques ; un jour ouvre sans heures
    apparait en 'Repos'."""
    sal = db.salarie(sid)
    semaines = semaines_du_mois(annee, mois)
    debut = semaines[0][0].isoformat()
    fin = semaines[-1][6].isoformat()
    data = db.jours_intervalle(sid, debut, fin)
    weeks = []
    total_mois = 0
    njours = 0
    for sem in semaines:
        jours = []
        wtot = 0
        for d in sem:
            if d.month != mois or d.year != annee:
                continue
            j = data.get(d.isoformat())
            mins = (heures_travaillees(j["amp_d"], j["amp_f"],
                                       j["pause_d"], j["pause_f"]) if j else 0)
            if mins == 0 and d.weekday() >= 5:
                continue  # week-end non travaille : masque
            jours.append({
                "date": d.strftime("%d/%m/%Y"),
                "jour": JOURS_NOMS[d.weekday()].capitalize(),
                "h": fmt_hm(mins) if mins > 0 else "Repos",
                "repos": mins == 0,
            })
            wtot += mins
            if mins > 0:
                njours += 1
        # On masque les semaines entierement non travaillees ; les semaines
        # affichees sont renumerotees sequentiellement a partir de 1.
        if jours and wtot > 0:
            weeks.append({"num": len(weeks) + 1, "jours": jours,
                          "total": fmt_hm(wtot)})
            total_mois += wtot
    return {"nom": nom_complet(sal), "duree": fmt_hm(db.duree_hebdo(sid)),
            "weeks": weeks, "total": fmt_hm(total_mois),
            "total_cent": fmt_centieme(total_mois), "njours": njours}


@app.route("/releve")
@admin_required
def releve():
    db = get_db()
    annee = _annee_param()
    mois = _mois_param()
    sals = db.salaries()
    selected_ids = _sids_selectionnes(db)
    # Page de selection si rien n'a encore ete demande explicitement.
    if not request.args.get("go"):
        return render_template("releve_select.html", sals=sals, annee=annee,
                               mois=mois, selection=set(selected_ids))
    rapports = [_releve_data(db, sid, annee, mois) for sid in selected_ids]
    return render_template("releve.html", rapports=rapports, annee=annee,
                           mois=mois, mois_nom=MOIS_NOMS[mois - 1],
                           edite=datetime.date.today().strftime("%d/%m/%Y"))


@app.route("/releve.xlsx")
@admin_required
def releve_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("Le module openpyxl n'est pas installé (pip install openpyxl).", "error")
        return redirect(url_for("releve"))
    db = get_db()
    annee = _annee_param()
    mois = _mois_param()
    selected_ids = _sids_selectionnes(db)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bleu = PatternFill("solid", fgColor="0B4F8A")
    jaune = PatternFill("solid", fgColor="FFD400")
    gris = PatternFill("solid", fgColor="EEF3F7")
    blanc_gras = Font(bold=True, color="FFFFFF")
    centre = Alignment(horizontal="center")
    bord = Border(*[Side(style="thin", color="C7D2DC")] * 4)

    for sid in selected_ids:
        r = _releve_data(db, sid, annee, mois)
        titre = (r["nom"] or "Salarié")[:28]
        ws = wb.create_sheet(title=titre or "Salarié")
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.merge_cells("A1:C1")
        ws["A1"] = "RELEVÉ MENSUEL DES HEURES TRAVAILLÉES"
        ws["A1"].font = Font(bold=True, size=13, color="C8102E")
        ws.merge_cells("A2:C2")
        ws["A2"] = f"{r['nom']}  —  {MOIS_NOMS[mois - 1]} {annee}"
        ws["A2"].font = Font(bold=True, size=11, color="0B4F8A")
        row = 4
        for w in r["weeks"]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c = ws.cell(row=row, column=1, value=f"Semaine {w['num']}")
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = bleu
            row += 1
            for col, txt in enumerate(["Jour", "Date", "Heures"], start=1):
                hc = ws.cell(row=row, column=col, value=txt)
                hc.font = blanc_gras
                hc.fill = PatternFill("solid", fgColor="1565C0")
                hc.alignment = centre
                hc.border = bord
            row += 1
            for j in w["jours"]:
                ws.cell(row=row, column=1, value=j["jour"]).border = bord
                ws.cell(row=row, column=2, value=j["date"]).border = bord
                hcell = ws.cell(row=row, column=3, value=j["h"])
                hcell.alignment = centre
                hcell.border = bord
                row += 1
            ws.cell(row=row, column=2, value="Total semaine").font = Font(bold=True)
            tc = ws.cell(row=row, column=3, value=w["total"])
            tc.font = Font(bold=True, color="1B7F3B")
            tc.alignment = centre
            tc.fill = gris
            row += 2
        ws.cell(row=row, column=1, value="TOTAL DU MOIS").font = Font(bold=True, size=12)
        tc = ws.cell(row=row, column=3, value=r["total"])
        tc.font = Font(bold=True, size=12, color="C8102E")
        tc.alignment = centre
        tc.fill = jaune

    if not wb.sheetnames:
        wb.create_sheet("Vide")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"releve_{MOIS_NOMS[mois - 1].lower()}_{annee}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument."
                 "spreadsheetml.sheet")


# --------------------------------------------------------------------------
# Salaries
# --------------------------------------------------------------------------
@app.route("/salaries", methods=["GET", "POST"])
@admin_required
def salaries():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "ajouter":
            nom = request.form.get("nom", "").strip()
            if nom:
                duree = parse_heure(request.form.get("duree", "")) or core.DUREE_HEBDO_DEFAUT
                db.ajouter_salarie(nom, request.form.get("prenom", "").strip(), duree)
                flash("Salarié ajouté.", "ok")
            else:
                flash("Le nom est obligatoire.", "error")
        elif action == "modifier":
            sid = int(request.form["sid"])
            duree = parse_heure(request.form.get("duree", "")) or core.DUREE_HEBDO_DEFAUT
            db.modifier_salarie(sid, request.form.get("nom", "").strip(),
                                request.form.get("prenom", "").strip(), duree)
            flash("Salarié modifié.", "ok")
        elif action == "supprimer":
            db.supprimer_salarie(int(request.form["sid"]))
            flash("Salarié supprimé.", "ok")
        return redirect(url_for("salaries"))

    sals = [{"id": s["id"], "nom": s["nom"], "prenom": s["prenom"] or "",
             "nom_complet": nom_complet(s), "duree": fmt_hm(db.duree_hebdo(s["id"]))}
            for s in db.salaries()]
    return render_template("salaries.html", sals=sals)


# --------------------------------------------------------------------------
# Comptes utilisateurs (admin)
# --------------------------------------------------------------------------
def _salarie_id_form():
    """Lit le salarie lie depuis le formulaire ('' -> None)."""
    v = request.form.get("salarie_id", "")
    try:
        return int(v) if v else None
    except (TypeError, ValueError):
        return None


@app.route("/comptes", methods=["GET", "POST"])
@admin_required
def comptes():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "ajouter":
            login_ = request.form.get("login", "").strip()
            pwd = request.form.get("password", "")
            if login_ and pwd:
                if db.get_user(login_):
                    flash("Cet identifiant existe déjà.", "error")
                else:
                    db.ajouter_user(login_, generate_password_hash(pwd),
                                    is_admin=1 if request.form.get("admin") else 0,
                                    salarie_id=_salarie_id_form())
                    flash("Compte créé.", "ok")
            else:
                flash("Identifiant et mot de passe obligatoires.", "error")
        elif action == "modifier":
            uid = int(request.form["uid"])
            db.modifier_user(uid, is_admin=1 if request.form.get("admin") else 0,
                             salarie_id=_salarie_id_form())
            flash("Compte mis à jour.", "ok")
        elif action == "motdepasse":
            login_ = request.form.get("login", "")
            pwd = request.form.get("password", "")
            if pwd:
                db.set_user_pwd(login_, generate_password_hash(pwd))
                flash("Mot de passe modifié.", "ok")
        elif action == "supprimer":
            uid = int(request.form["uid"])
            if len(db.users()) <= 1:
                flash("Impossible de supprimer le dernier compte.", "error")
            else:
                db.supprimer_user(uid)
                flash("Compte supprimé.", "ok")
        return redirect(url_for("comptes"))

    sals = db.salaries()
    noms = {s["id"]: nom_complet(s) for s in sals}
    users = []
    nb_admins = 0
    for u in db.users():
        if u["is_admin"]:
            nb_admins += 1
        users.append({
            "id": u["id"], "login": u["login"], "is_admin": bool(u["is_admin"]),
            "salarie_id": u["salarie_id"],
            "salarie_nom": noms.get(u["salarie_id"], "—"),
        })
    return render_template("comptes.html", users=users, sals=sals,
                           noms=noms, nb_admins=nb_admins)


# --------------------------------------------------------------------------
# Espace salarié (vue simplifiée : saisir SA journée)
# --------------------------------------------------------------------------
def _pause_totale(paires):
    """Somme des durées de pause (minutes) à partir de paires (debut, fin)."""
    tot = 0
    for d, f in paires:
        if d is not None and f is not None and f > d:
            tot += f - d
    return tot


@app.route("/espace", methods=["GET", "POST"])
@login_required
def espace():
    db = get_db()
    sid = session.get("salarie_id")
    login_ = session.get("user")
    today = datetime.date.today()
    iso = today.isoformat()

    sal = db.salarie(sid) if sid else None
    if not sal:
        return render_template("espace.html", sal=None, login=login_)

    if request.method == "POST":
        # Securite : un salarie ne saisit que SA journee du jour meme.
        if request.form.get("date") != iso:
            flash("Vous ne pouvez saisir que la journée d'aujourd'hui.", "error")
            return redirect(url_for("espace"))
        try:
            amp_d = parse_heure(request.form.get("amp_d", ""))
            amp_f = parse_heure(request.form.get("amp_f", ""))
        except ValueError:
            amp_d = amp_f = None
        if amp_d is None or amp_f is None or amp_f <= amp_d:
            flash("Indiquez une heure d'arrivée et une heure de départ valides "
                  "(départ après l'arrivée).", "error")
            return redirect(url_for("espace"))
        # pauses multiples
        paires = []
        debuts = request.form.getlist("pause_d")
        fins = request.form.getlist("pause_f")
        for pd_, pf_ in zip(debuts, fins):
            try:
                a = parse_heure(pd_)
                b = parse_heure(pf_)
            except ValueError:
                continue
            if a is not None and b is not None and b > a:
                paires.append((a, b))
        ptot = _pause_totale(paires)
        # Encodage : une seule pause reelle -> on garde ses horaires ;
        # plusieurs pauses -> total encode en (0, total) (calculs inchangés).
        if len(paires) == 1:
            pd_db, pf_db = paires[0]
        elif ptot > 0:
            pd_db, pf_db = 0, ptot
        else:
            pd_db, pf_db = None, None
        db.enregistrer_jour(sid, iso, {
            "amp_d": amp_d, "amp_f": amp_f, "pause_d": pd_db, "pause_f": pf_db,
            "km_jour": None, "km_com": None, "km_vide": None})
        db.commit()
        db.set_pauses_jour(sid, iso, paires)
        flash("Journée enregistrée.", "ok")
        return redirect(url_for("espace"))

    # --- GET : preparation de l'affichage ---
    jour = db.jour(sid, iso)
    pauses = [{"d": fmt_hm(p["debut"]), "f": fmt_hm(p["fin"])}
              for p in db.pauses_jour(sid, iso)]
    if jour and not pauses and jour["pause_d"] is not None and jour["pause_f"] is not None:
        pauses = [{"d": fmt_hm(jour["pause_d"]), "f": fmt_hm(jour["pause_f"])}]
    saisie_jour = None
    if jour:
        saisie_jour = {
            "amp_d": fmt_hm(jour["amp_d"]) if jour["amp_d"] is not None else "",
            "amp_f": fmt_hm(jour["amp_f"]) if jour["amp_f"] is not None else "",
            "travail": fmt_hm(heures_travaillees(
                jour["amp_d"], jour["amp_f"], jour["pause_d"], jour["pause_f"])),
        }

    # historique du mois en cours (lecture seule)
    debut = today.replace(day=1).isoformat()
    data = db.jours_intervalle(sid, debut, iso)
    histo = []
    for d_iso in sorted(data.keys(), reverse=True):
        j = data[d_iso]
        d = datetime.date.fromisoformat(d_iso)
        histo.append({
            "date": JOURS_NOMS[d.weekday()][:3].lower() + " " + d.strftime("%d/%m"),
            "amp": (fmt_hm(j["amp_d"]) + " → " + fmt_hm(j["amp_f"])
                    if j["amp_d"] is not None and j["amp_f"] is not None else "—"),
            "travail": fmt_hm(heures_travaillees(
                j["amp_d"], j["amp_f"], j["pause_d"], j["pause_f"])),
            "aujourdhui": d_iso == iso,
        })

    return render_template(
        "espace.html", sal=sal, login=login_, nom=nom_complet(sal),
        date_iso=iso,
        date_fr=(JOURS_NOMS[today.weekday()].lower() + " " + str(today.day) + " "
                 + MOIS_NOMS[today.month - 1].lower() + " " + str(today.year)),
        saisie=saisie_jour, pauses=pauses, histo=histo)


# --------------------------------------------------------------------------
# Export Excel
# --------------------------------------------------------------------------
@app.route("/export.xlsx")
@admin_required
def export_excel():
    try:
        import openpyxl
    except ImportError:
        flash("Le module openpyxl n'est pas installé (pip install openpyxl).", "error")
        return redirect(url_for("recap_global"))
    db = get_db()
    annee = _annee_param()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RECAPITULATIF"
    entete = ["SALARIÉ", "DURÉE HEBDO", "MOIS", "Total", "Sup>35h",
              "Dans accord", "Au-delà (acquis)", "Payées", "Congés pris",
              "Solde compteur"]
    ws.append(entete)
    for s in db.salaries():
        for mois in range(1, 13):
            t = totaux_mois(db, s["id"], annee, mois)
            ws.append([
                nom_complet(s), fmt_hm(db.duree_hebdo(s["id"])), MOIS_NOMS[mois - 1],
                fmt_hm(t["total"]), fmt_hm(t["sup35"]), fmt_hm(t["accord39"]),
                fmt_hm(t["sup39"]), fmt_hm(db.payees(s["id"], annee, mois)),
                fmt_hm(db.conges_pris(s["id"], annee, mois)),
                fmt_hm_signe(db.solde_conges(s["id"], annee, mois)),
            ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"heures_{annee}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument."
                              "spreadsheetml.sheet")


if __name__ == "__main__":
    # HOST : 127.0.0.1 par defaut (acces local au PC uniquement).
    #   - reseau local (Wi-Fi du bureau) : HOST=0.0.0.0
    #   - en ligne : on utilise plutot gunicorn (voir Procfile / DEPLOIEMENT.md)
    port = int(os.environ.get("PORT", 5000))
    host = os.environ.get("HOST", "127.0.0.1")
    app.run(host=host, port=port, debug=False)
